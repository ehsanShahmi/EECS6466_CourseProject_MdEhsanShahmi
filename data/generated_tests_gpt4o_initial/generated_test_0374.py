import unittest
import os
import tempfile
from openpyxl import Workbook
from your_module import task_func  # Adjust the import based on your module structure

class TestTaskFunc(unittest.TestCase):

    def setUp(self):
        """Create a temporary directory and test files before each test."""
        self.test_dir = tempfile.mkdtemp()
        self.file_paths = []

    def tearDown(self):
        """Remove the temporary directory after each test."""
        for file_path in self.file_paths:
            os.remove(file_path)
        os.rmdir(self.test_dir)

    def create_test_file(self, filename, content):
        """Helper method to create an Excel file with specified content."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(content)
        file_path = os.path.join(self.test_dir, filename)
        workbook.save(file_path)
        self.file_paths.append(file_path)

    def test_no_files(self):
        """Test case for empty directory, expecting zero processed files."""
        result = task_func(self.test_dir)
        self.assertEqual(result, 0)

    def test_one_file_no_quotes(self):
        """Test case where Excel file has no double quotes, expecting one processed file."""
        self.create_test_file('test_no_quotes.xlsx', ['This is a test string.'])
        result = task_func(self.test_dir)
        self.assertEqual(result, 1)

    def test_one_file_with_quotes(self):
        """Test case where Excel file has double quotes, expecting quotes to be escaped."""
        self.create_test_file('test_with_quotes.xlsx', ['This is a "test" string.'])
        task_func(self.test_dir)
        workbook = load_workbook(os.path.join(self.test_dir, 'test_with_quotes.xlsx'))
        cell_value = workbook.active.cell(row=1, column=1).value
        expected_value = 'This is a \\"test\\" string.'
        self.assertEqual(cell_value, expected_value)

    def test_multiple_files(self):
        """Test case with multiple Excel files with and without quotes."""
        self.create_test_file('file1.xlsx', ['No quotes here.'])
        self.create_test_file('file2.xlsx', ['Quotes "here" and "there".'])
        result = task_func(self.test_dir)
        self.assertEqual(result, 2)

    def test_invalid_directory(self):
        """Test case for invalid directory, expecting a FileNotFoundError."""
        with self.assertRaises(FileNotFoundError):
            task_func('/invalid/path/')

if __name__ == '__main__':
    unittest.main()