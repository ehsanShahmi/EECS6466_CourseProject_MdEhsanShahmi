import os
import unittest

# Here is your prompt:
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    """
    Converts an Excel file (.xls or .xlsx) to a CSV file by reading the contents of the Excel file
    and writing them to a new CSV file with the same name but a different extension. Allows specifying
    separate paths for the Excel file source and the CSV file destination.

    Parameters:
        file_name (str): The name of the Excel file to be converted.
        excel_file_path (str): The directory path where the Excel file is located.
        csv_file_path (str): The directory path where the CSV file should be saved.

    Returns:
        str: The name of the created CSV file.

    Requirements:
    - openpyxl.load_workbook
    - os
    - csv

    Example:
    >>> task_func('test.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    'test.csv'
    >>> task_func('nonexistent.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    Traceback (most recent call last):
       ...
    FileNotFoundError: [Errno 2] No such file or directory: '/path/to/excel/files/nonexistent.xlsx'

    Note:
    - This function assumes the active sheet is the one to be converted.
    """

class TestTaskFunc(unittest.TestCase):

    def setUp(self):
        # Create temporary directories for testing
        os.makedirs('test_excel', exist_ok=True)
        os.makedirs('test_csv', exist_ok=True)

        # Create a sample Excel file for testing
        self.test_excel_file = 'test_excel/test.xlsx'
        self.create_test_excel(self.test_excel_file, [['Header1', 'Header2'], ['Value1', 'Value2']])

    def tearDown(self):
        # Cleanup temporary files and directories
        try:
            os.remove(self.test_excel_file)
        except FileNotFoundError:
            pass
        try:
            os.remove('test_csv/test.csv')
        except FileNotFoundError:
            pass
        os.rmdir('test_excel')
        os.rmdir('test_csv')

    def create_test_excel(self, file_path, data):
        """Creates a sample Excel file for testing purposes."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(file_path)

    def test_converts_excel_to_csv_successfully(self):
        result = task_func('test.xlsx', 'test_excel', 'test_csv')
        self.assertEqual(result, 'test.csv')
        with open('test_csv/test.csv', 'r', encoding='utf-8') as f:
            content = f.readlines()
            self.assertEqual(content[0].strip(), 'Header1,Header2')
            self.assertEqual(content[1].strip(), 'Value1,Value2')

    def test_file_not_found_exception(self):
        with self.assertRaises(FileNotFoundError):
            task_func('nonexistent.xlsx', 'test_excel', 'test_csv')

    def test_invalid_excel_file_format(self):
        self.create_invalid_test_excel('test_excel/invalid_file.txt', 'This is not a valid excel file.')
        with self.assertRaises(Exception):  # Replace with specific exception related to invalid formatting
            task_func('invalid_file.txt', 'test_excel', 'test_csv')

    def create_invalid_test_excel(self, file_path, content):
        """Creates an invalid file for testing purposes."""
        with open(file_path, 'w') as f:
            f.write(content)

    def test_empty_excel_file(self):
        self.create_test_excel('test_excel/empty_test.xlsx', [])
        result = task_func('empty_test.xlsx', 'test_excel', 'test_csv')
        self.assertEqual(result, 'empty_test.csv')
        with open('test_csv/empty_test.csv', 'r', encoding='utf-8') as f:
            content = f.readlines()
            self.assertEqual(content, [])

if __name__ == '__main__':
    unittest.main()